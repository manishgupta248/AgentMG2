"""
Excel write tools.

IMPORTANT TRADEOFF: openpyxl's read_only=True mode is read-only by
design and cannot write. Writing requires a normal load (keep_vba=False,
read_only=False), which loads more into memory than the streaming reads
in excel_read.py. This is acceptable for typical personal-use workbook
sizes but is NOT streaming — a very large file (tens of MB+) being
written to could pressure the 8GB RAM budget. If that becomes a real
need, revisit with openpyxl's write-only mode (which streams writes
but cannot read/modify existing content in the same pass).
"""

from __future__ import annotations

from pathlib import Path

from pydantic import BaseModel, Field
import openpyxl

from app.core.registry import tool
from app.core.types import PermissionLevel
from app.core.exceptions import ToolExecutionError


def _open_workbook_writable(file_path: str):
    path = Path(file_path)
    if not path.exists():
        raise ToolExecutionError(f"Excel file not found: {file_path}", context={"file_path": file_path})
    try:
        return openpyxl.load_workbook(filename=str(path), read_only=False, data_only=False)
    except Exception as e:
        raise ToolExecutionError(f"Failed to open Excel file for writing: {e}", context={"file_path": file_path}) from e


class WriteRangeInput(BaseModel):
    file_path: str
    sheet_name: str | None = None
    start_cell: str = Field(..., description="Top-left cell to start writing at, e.g. 'A1'")
    rows: list[list] = Field(..., description="2D list of row values to write, starting at start_cell")
    model_config = {"extra": "forbid"}


@tool(
    "excel_write_range",
    permission=PermissionLevel.MODIFY,
    description="Writes a 2D block of values into an Excel sheet starting at a given cell.",
    input_schema=WriteRangeInput,
    example_phrases=["write this data to the excel file", "update the spreadsheet with these values"],
)

def excel_write_range(file_path: str, start_cell: str, rows: list, sheet_name: str | None = None) -> dict:
    wb = _open_workbook_writable(file_path)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
        for r_offset, row_values in enumerate(rows):
            for c_offset, value in enumerate(row_values):
                ws.cell(row=start_row + r_offset, column=start_col + c_offset, value=value)
        wb.save(file_path)
        return {"sheet": ws.title, "cells_written": sum(len(r) for r in rows), "start_cell": start_cell}
    except KeyError as e:
        raise ToolExecutionError(f"Sheet not found: {sheet_name}", context={"file_path": file_path}) from e
    finally:
        wb.close()


class AppendRowsInput(BaseModel):
    file_path: str
    sheet_name: str | None = None
    rows: list[list] = Field(..., description="Rows to append after the last existing row")
    model_config = {"extra": "forbid"}


@tool(
    "excel_append_rows",
    permission=PermissionLevel.MODIFY,
    description="Appends rows to the end of an Excel sheet.",
    input_schema=AppendRowsInput,
    example_phrases=["add a row to the spreadsheet", "append this data to the excel file"],
)
def excel_append_rows(file_path: str, rows: list, sheet_name: str | None = None) -> dict:
    wb = _open_workbook_writable(file_path)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        for row_values in rows:
            ws.append(row_values)
        wb.save(file_path)
        return {"sheet": ws.title, "rows_appended": len(rows), "new_max_row": ws.max_row}
    except KeyError as e:
        raise ToolExecutionError(f"Sheet not found: {sheet_name}", context={"file_path": file_path}) from e
    finally:
        wb.close()