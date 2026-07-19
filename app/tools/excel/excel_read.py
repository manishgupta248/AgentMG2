"""
Excel read tools — openpyxl in read_only=True streaming mode throughout,
per the project's 8GB RAM discipline. Never loads a full workbook into
memory; iterates rows lazily.
"""

from __future__ import annotations

from pathlib import Path

from pydantic import BaseModel, Field
import openpyxl

from app.core.registry import tool
from app.core.types import PermissionLevel
from app.core.exceptions import ToolExecutionError


def _open_workbook_readonly(file_path: str):
    path = Path(file_path)
    if not path.exists():
        raise ToolExecutionError(f"Excel file not found: {file_path}", context={"file_path": file_path})
    try:
        return openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as e:
        raise ToolExecutionError(f"Failed to open Excel file: {e}", context={"file_path": file_path}) from e


class ListSheetsInput(BaseModel):
    file_path: str
    model_config = {"extra": "forbid"}


@tool(
    "excel_list_sheets",
    permission=PermissionLevel.READ,
    description="Lists all sheet names in an Excel workbook.",
    input_schema=ListSheetsInput,
    example_phrases=["list sheets in this excel file", "what sheets are in the workbook"],
)
def excel_list_sheets(file_path: str) -> dict:
    wb = _open_workbook_readonly(file_path)
    try:
        return {"sheets": wb.sheetnames}
    finally:
        wb.close()


class ReadRangeInput(BaseModel):
    file_path: str
    sheet_name: str | None = Field(None, description="If omitted, uses the active sheet")
    cell_range: str = Field(..., description="e.g. 'A1:C10'")
    model_config = {"extra": "forbid"}


@tool(
    "excel_read_range",
    permission=PermissionLevel.READ,
    description="Reads a specific cell range (e.g. A1:C10) from an Excel sheet.",
    input_schema=ReadRangeInput,
    example_phrases=["read cells A1 to C10", "get the range from the spreadsheet"],
)
def excel_read_range(file_path: str, cell_range: str, sheet_name: str | None = None) -> dict:
    wb = _open_workbook_readonly(file_path)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = []
        for row in ws[cell_range]:
            rows.append([cell.value for cell in row])
        return {"sheet": ws.title, "range": cell_range, "rows": rows}
    except KeyError as e:
        raise ToolExecutionError(f"Sheet not found: {sheet_name}", context={"file_path": file_path}) from e
    finally:
        wb.close()


class ReadSheetInput(BaseModel):
    file_path: str
    sheet_name: str | None = None
    max_rows: int = Field(1000, description="Safety cap to avoid loading huge sheets fully into memory/response")
    model_config = {"extra": "forbid"}


@tool(
    "excel_read_sheet",
    permission=PermissionLevel.READ,
    description="Reads all rows from an Excel sheet, streaming (capped at max_rows for safety).",
    input_schema=ReadSheetInput,
    example_phrases=["read the whole sheet", "show me all the data in this excel file"],
)
def excel_read_sheet(file_path: str, sheet_name: str | None = None, max_rows: int = 1000) -> dict:
    wb = _open_workbook_readonly(file_path)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = []
        truncated = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                truncated = True
                break
            rows.append(list(row))
        return {"sheet": ws.title, "row_count": len(rows), "truncated": truncated, "rows": rows}
    except KeyError as e:
        raise ToolExecutionError(f"Sheet not found: {sheet_name}", context={"file_path": file_path}) from e
    finally:
        wb.close()